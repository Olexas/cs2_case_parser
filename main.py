import requests
from lxml import html, etree
from openpyxl import Workbook, load_workbook
import sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.proxy import StyleProxy

CONST_KEY_PRICE = 230
CONST_GOLD_PROB = 0.0025575
CONST_RED_PROB = 0.0063939
CONST_PINK_PROB = 0.0319639
CONST_PURPLE_PROB = 0.1598465
CONST_BLUE_PROB = 0.7992327
CONST_FN = 0.07
CONST_MW = 0.08
CONST_FT = 0.23
CONST_WW = 0.07
CONST_BS = 0.55

skip_flag = False


def read_file(filename):
    with open(filename, encoding="utf-8") as input_file:
        text = input_file.read()
    return text


def getting_chances(gold, red, pink, purple, blue):
    if blue:
        if gold:
            return CONST_GOLD_PROB, CONST_RED_PROB, CONST_PINK_PROB, CONST_PURPLE_PROB, CONST_BLUE_PROB
        if red:
            if pink:
                return 0, CONST_RED_PROB, CONST_PINK_PROB, CONST_PURPLE_PROB, CONST_BLUE_PROB + CONST_GOLD_PROB
            return 0, CONST_RED_PROB, 0, CONST_PURPLE_PROB, CONST_BLUE_PROB + CONST_GOLD_PROB + CONST_PINK_PROB
        if pink:
            if purple:
                return 0, 0, CONST_PINK_PROB, CONST_PURPLE_PROB, CONST_BLUE_PROB + CONST_GOLD_PROB + CONST_RED_PROB
            return 0, 0, CONST_PINK_PROB, 0, CONST_BLUE_PROB + CONST_GOLD_PROB + CONST_RED_PROB + CONST_PURPLE_PROB
        if purple:
            return 0, 0, 0, CONST_PURPLE_PROB, CONST_BLUE_PROB + CONST_GOLD_PROB + CONST_RED_PROB + CONST_PINK_PROB
        return 0, 0, 0, 0, 1

    if purple:
        if pink:
            return 0, 0, CONST_PURPLE_PROB, CONST_BLUE_PROB + CONST_GOLD_PROB + CONST_RED_PROB + CONST_PINK_PROB, 0
        return 0, 0, 0, 1, 0

    return 0, 0, 1, 0, 0


def getting_float(float_range):
    float_min = float_range[0]
    float_max = 1 - float_range[1]

    float_full = 1 - (float_max + float_min)

    float_fn, float_min = max(CONST_FN - float_min, 0), max(float_min - CONST_FN, 0)
    float_mw, float_min = max(CONST_MW - float_min, 0), max(float_min - CONST_MW, 0)
    float_ft, float_min = max(CONST_FT - float_min, 0), max(float_min - CONST_FT, 0)
    float_ww, float_min = max(CONST_WW - float_min, 0), max(float_min - CONST_WW, 0)
    float_bs, float_min = max(CONST_BS - float_min, 0), max(float_min - CONST_BS, 0)

    float_bs, float_max = max(float_bs - float_max, 0), max(float_max - CONST_BS, 0)
    float_ww, float_max = max(float_ww - float_max, 0), max(float_max - CONST_WW, 0)
    float_ft, float_max = max(float_ft - float_max, 0), max(float_max - CONST_FT, 0)
    float_mw, float_max = max(float_mw - float_max, 0), max(float_max - CONST_MW, 0)
    float_fn, float_max = max(float_fn - float_max, 0), max(float_max - CONST_FN, 0)

    if round(
            float_fn / float_full + float_mw / float_full + float_ft / float_full + float_ww / float_full \
            + float_bs / float_full,
            6) != 1 and float_fn / float_full + float_mw / float_full + float_ft / float_full \
            + float_ww / float_full + float_bs / float_full != 1.0000000000000002:
        print(
            'ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR')
        print(
            float_fn / float_full + float_mw / float_full + float_ft / float_full + float_ww / float_full
            + float_bs / float_full)
        a = input()

    print('Вероятности качеств:', float_fn / float_full, float_mw / float_full,
          float_ft / float_full,
          float_ww / float_full, float_bs / float_full)

    return float_fn / float_full, float_mw / float_full, float_ft / float_full, \
           float_ww / float_full, float_bs / float_full


def item_load(item, cell_x, cell_y, item_rarity, case_price, items_info, page, color,
              dispersion_list, key_price):
    item_name = f"{item.xpath('.//h3/a/text()')[0]} | " \
                f"{item.xpath('.//h3/a/text()')[1]} ({round(item_rarity * 100, 2)}%)"

    item_link = item.xpath('.//a/@href')[3]

    r = requests.get(item_link)

    with open('item_page.html', 'w', encoding="utf-8") as output_file:
        output_file.write(r.text)

    item_page_html = read_file('item_page.html')

    item_tree = html.fromstring(item_page_html)

    prices_steam_links = item_tree.xpath('//div[@class = "tab-pane active"]/'
                                         'div[@class = "btn-group-sm btn-group-justified"]/a/@href')
    prices_steam_links.reverse()

    prices = item_tree.xpath('//div[@class = "tab-pane active"]/'
                             'div[@class = "btn-group-sm btn-group-justified"]/a/'
                             'span[@class = "pull-right"]/text()')

    vars = ['st_fn', 'st_mw', 'st_ft', 'st_ww', 'st_bs', 'fn', 'mw', 'ft', 'ww', 'bs']

    vars.reverse()

    for price_index in range(len(prices)):
        if prices[price_index] == 'No Recent Price':
            pass

    for price_index in range(len(prices)):
        if prices[price_index] != 'Not Possible' and prices[
            price_index] != 'No Recent Price':
            prices[price_index] = float(
                prices[price_index].replace(' pуб.', '').replace(',', '.').replace(' ',
                                                                                   ''))
        elif prices[price_index] == 'Not Possible':
            prices[price_index] = 0
        else:
            prices[price_index] = 'No Recent Price'

    prices.reverse()

    global skip_flag

    price_index_for_steam_links = 0
    for price_index in range(len(prices)):
        if prices[price_index] == 0:
            continue
        if 'No Recent Price' in prices:
            if prices[price_index] == 'No Recent Price':
                while True:
                    if not skip_flag:
                        print(prices_steam_links[price_index_for_steam_links])
                        price_input = input(
                            f'На данный момент цена для ({item_name} {vars[price_index]}) недоступна. '
                            f'Введите ее вручную или напишите "skip" для автоподбора цены: ')

                        if price_input == 'skip':
                            prices[price_index] = max(
                                list(filter(lambda x: x != 'No Recent Price', prices)))
                            skip_flag = True
                            break

                        else:
                            try:
                                prices[price_index] = float(price_input)
                                break

                            except Exception:
                                print(
                                    'Ввод не является ценой предмета, попробуйте еще раз')
                    else:
                        prices[price_index] = max(
                            list(filter(lambda x: x != 'No Recent Price', prices)))
                        break

            price_index_for_steam_links += 1

    prices.reverse()
    vars.reverse()

    payback_qualities_list = list()
    st_payback_qualities_list = list()

    if len(prices) == 10:
        st_fn_price, st_mw_price, st_ft_price, st_ww_price, st_bs_price = \
            prices[0], prices[1], prices[2], prices[3], prices[4]
        fn_price, mw_price, ft_price, ww_price, bs_price = \
            prices[5], prices[6], prices[7], prices[8], prices[9]

        print(f'|||{item_name}|||')
        print()

        print('цены:', prices, case_price)

        float_range = [float(i) for i in
                       item_tree.xpath('//div[@class = "marker-value cursor-default"]'
                                       '/text()')]
        print('float range:', float_range)
        floats = getting_float(float_range)

        dds = []

        for price_index in range(len(prices)):
            if price_index in range(0, 5):
                full_item_rarity = floats[price_index % 5] * item_rarity * 0.1
            else:
                full_item_rarity = floats[price_index % 5] * item_rarity * 0.9
            dispersion_list.append([prices[price_index], full_item_rarity])
            dds.append([prices[price_index], full_item_rarity])

        print('дисперсия:', dds)

        float_fn_price, float_mw_price, float_ft_price, float_ww_price, float_bs_price = \
            fn_price * floats[0], mw_price * floats[1], ft_price * floats[2], ww_price * \
            floats[3], \
            bs_price * floats[4]
        print('цены с учетом флоата:', float_fn_price, float_mw_price, float_ft_price,
              float_ww_price, float_bs_price)

        average_item_price = \
            float_bs_price + float_fn_price + float_ft_price + float_mw_price + float_ww_price
        print('средняя цена без st:', average_item_price)
        print()

        float_st_fn_price, float_st_mw_price, float_st_ft_price, float_st_ww_price, float_st_bs_price = \
            st_fn_price * floats[0], st_mw_price * floats[1], st_ft_price * floats[2], \
            st_ww_price * floats[3], st_bs_price * floats[4]
        print('цены с учетом флоата st:', float_st_fn_price, float_st_mw_price,
              float_st_ft_price, float_st_ww_price,
              float_st_bs_price)

        average_st_item_price = \
            float_st_bs_price + float_st_fn_price + float_st_ft_price + float_st_mw_price + \
            float_st_ww_price
        print('средняя цена st:', average_st_item_price)
        print()

        item_price = average_item_price * 0.9 + average_st_item_price * 0.1
        print('цена скина:', item_price)
        print()

        st_payback_count = 0
        st_var_count = 0
        payback_count = 0
        var_count = 0

        for price_index in range(len(prices)):
            if prices[price_index] == 0:
                continue
            if price_index in range(0, 5):
                if prices[price_index] > key_price + case_price:
                    st_payback_count += 1
                    st_payback_qualities_list.append(vars[price_index])
                st_var_count += 1
            else:
                if prices[price_index] > key_price + case_price:
                    payback_count += 1
                    payback_qualities_list.append(vars[price_index])
                var_count += 1

        average_item_payback = item_price / (key_price + case_price)

        print('все остальное:',
              [payback_count, var_count, st_payback_count, st_var_count, item_price,
               average_item_payback])
        print()

        excel_decoration(page, cell_x, cell_y, 17)
        page.cell(column=cell_x, row=cell_y, value=item_name).font = Font(color=color)
        cell_x += 1

        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(fn_price, 2)} руб.').font = Font(color=color)
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[0] * 100, 2)}%').font = Font(color=color)
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(mw_price, 2)} руб.').font = Font(color=color)
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[1] * 100, 2)}%').font = Font(color=color)
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(ft_price, 2)} руб.').font = Font(color=color)
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[2] * 100, 2)}%').font = Font(color=color)
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(ww_price, 2)} руб.').font = Font(color=color)
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[3] * 100, 2)}%').font = Font(color=color)
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(bs_price, 2)} руб.').font = Font(color=color)
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[4] * 100, 2)}%').font = Font(color=color)
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(st_fn_price, 2)} руб.').font = Font(color='ed7d31')
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[0] * 100, 2)}%').font = Font(color='ed7d31')
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(st_mw_price, 2)} руб.').font = Font(color='ed7d31')
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[1] * 100, 2)}%').font = Font(color='ed7d31')
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(st_ft_price, 2)} руб.').font = Font(color='ed7d31')
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[2] * 100, 2)}%').font = Font(color='ed7d31')
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(st_ww_price, 2)} руб.').font = Font(color='ed7d31')
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[3] * 100, 2)}%').font = Font(color='ed7d31')
        cell_y -= 1
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(st_bs_price, 2)} руб.').font = Font(color='ed7d31')
        cell_y += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(floats[4] * 100, 2)}%').font = Font(color='ed7d31')
        cell_y -= 1
        cell_x += 1

        page.cell(column=cell_x, row=cell_y,
                  value=f'{payback_count}/{var_count}').font = Font(color=color)
        if len(payback_qualities_list) > 0:
            page.cell(column=cell_x, row=cell_y + 1,
                      value=f'({", ".join(payback_qualities_list)})').font = \
                Font(color=color)
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{st_payback_count}/{st_var_count}').font = Font(color='ed7d31')
        if len(st_payback_qualities_list) > 0:
            page.cell(column=cell_x, row=cell_y + 1,
                      value=f'({", ".join(st_payback_qualities_list).replace("st_", "")}'
                            f')').font = Font(color='ed7d31')
        cell_x += 1

        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(average_item_price, 2)} руб.').font = Font(color=color)
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(average_st_item_price, 2)} руб.').font = Font(
            color='ed7d31')
        cell_x += 1
        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(item_price, 2)} руб.').font = Font(color=color)
        cell_x += 1

        page.cell(column=cell_x, row=cell_y,
                  value=f'{round(average_item_payback * 100, 2)}%').font = Font(
            color=color)

        items_info.append([payback_count, var_count, st_payback_count, st_var_count,
                           average_item_price,
                           average_st_item_price, item_price, average_item_payback])


def load_info_for_case(items_info, page, cell_x, cell_y, case_info, color, item_rarity):
    payback_variants = 0
    item_variants = 0
    st_payback_variants = 0
    st_item_variants = 0
    average_items_price = 0
    average_st_items_price = 0
    average_full_price = 0
    average_payback_present = 0

    final_score_for_rarity = 0
    average_skins_cost_for_dispersion = 0

    excel_decoration(page, cell_x, cell_y, 17, True)

    for item_info in items_info:
        payback_variants += item_info[0]
        item_variants += item_info[1]
        st_payback_variants += item_info[2]
        st_item_variants += item_info[3]
        average_items_price += item_info[4]
        average_st_items_price += item_info[5]
        average_full_price += item_info[6]
        average_payback_present += item_info[7]

        final_score_for_rarity += item_info[7] * item_rarity

        average_skins_cost_for_dispersion += item_info[6] * item_rarity

    variants_payback = payback_variants / item_variants
    st_variants_payback = st_payback_variants / st_item_variants
    average_items_price = round(average_items_price / len(items_info), 2)
    average_st_items_price = round(average_st_items_price / len(items_info), 2)
    average_full_price = round(average_full_price / len(items_info), 2)
    average_payback_present = average_payback_present / len(items_info)

    case_info.append([variants_payback, st_variants_payback, average_payback_present,
                      final_score_for_rarity,
                      average_skins_cost_for_dispersion])

    page.cell(column=14, row=cell_y,
              value=f'{payback_variants}/{item_variants} ({round(variants_payback * 100, 2)}%)').font = \
        Font(color=color)
    page.cell(column=15, row=cell_y,
              value=f'{st_payback_variants}/{st_item_variants} ({round(st_variants_payback * 100, 2)}%)').font = \
        Font(color='ed7d31')
    page.cell(column=16, row=cell_y,
              value=f'{average_items_price} руб.').font = \
        Font(color=color)
    page.cell(column=17, row=cell_y,
              value=f'{average_st_items_price} руб.').font = \
        Font(color='ed7d31')
    page.cell(column=18, row=cell_y,
              value=f'{average_full_price} руб.').font = \
        Font(color=color)
    page.cell(column=19, row=cell_y,
              value=f'{round(average_payback_present * 100, 2)}%.').font = \
        Font(color=color)


def excel_decoration_for_rarity(page, cell_x, cell_y, length):
    thins = Side(border_style="thin", color="000000")
    if length == 17:
        for cell_num in range(1, length + 1):
            if cell_num in range(2, 12):
                if cell_num == 6:
                    page.cell(column=cell_x, row=cell_y, value='').border = Border(
                        bottom=thins, right=thins)
                else:
                    page.cell(column=cell_x, row=cell_y, value='').border = Border(
                        bottom=thins)
            else:
                page.cell(column=cell_x, row=cell_y, value='').border = Border(
                    bottom=thins, right=thins, left=thins)
            cell_x += 1


def excel_decoration(page, cell_x, cell_y, length, last_flag=False):
    thins = Side(border_style="thin", color="000000")
    double = Side(border_style="medium", color="000000")
    if length == 17:
        for cell_num in range(1, length + 1):
            if cell_num in range(2, 12):
                if cell_num == 6:
                    page.cell(column=cell_x, row=cell_y, value='').border = Border(
                        right=thins)
            else:
                page.cell(column=cell_x, row=cell_y, value='').border = Border(
                    right=thins, left=thins)
                if cell_num >= 12:
                    page.cell(column=cell_x, row=cell_y, value='').alignment = Alignment(
                        horizontal='right')
                    page.cell(column=cell_x, row=cell_y + 1,
                              value='').alignment = Alignment(horizontal='right')
                    if last_flag:
                        page.cell(column=cell_x, row=cell_y, value='').border = Border(
                            right=double, left=double,
                            top=double)
            cell_x += 1
        cell_y += 1
        cell_x = 3
        for cell_num in range(1, length + 1):
            if cell_num in range(2, 12):
                if cell_num == 6:
                    page.cell(column=cell_x, row=cell_y, value='').border = Border(
                        bottom=thins, right=thins)
                else:
                    page.cell(column=cell_x, row=cell_y, value='').border = Border(
                        bottom=thins)
            else:
                page.cell(column=cell_x, row=cell_y, value='').border = Border(
                    bottom=thins, right=thins, left=thins)
                if cell_num >= 12 and last_flag:
                    page.cell(column=cell_x, row=cell_y, value='').border = Border(
                        right=double, left=double,
                        bottom=double)
            cell_x += 1


def dispersion_score(dispersion_list, average_dispersion_cost_for_case, case_price):
    dispersion_case_score = 0
    print(average_dispersion_cost_for_case)
    for dispersion_item_info in dispersion_list:
        item_value = ((dispersion_item_info[
                           0] - average_dispersion_cost_for_case) ** 2) * \
                     dispersion_item_info[1]
        dispersion_case_score += item_value
    print(f'!!!{dispersion_case_score}!!!')
    return dispersion_case_score / average_dispersion_cost_for_case


def case_load(case_name, case_link, case_price, page, cell_x, cell_y, case_num,
              key_flag=True):
    r = requests.get(case_link, params={'sort': 'price_steam_agg', 'sort_agg': 'avg',
                                        'order': 'asc'})

    with open('case_page.html', 'w', encoding="utf-8") as output_file:
        output_file.write(r.text)

    case_page_html = read_file('case_page.html')

    tree = html.fromstring(case_page_html)

    gold_items = tree.xpath('//div[@class = "quality color-rare-item"]')
    red_items = tree.xpath('//div[@class = "quality color-covert"]')
    pink_items = tree.xpath('//div[@class = "quality color-classified"]')
    purple_items = tree.xpath('//div[@class = "quality color-restricted"]')
    blue_items = tree.xpath('//div[@class = "quality color-milspec"]')

    all_items = tree.xpath('//div[@class = "well result-box nomargin"]')

    global gold_rarity, red_rarity, pink_rarity, purple_rarity, blue_rarity
    gold_rarity, red_rarity, pink_rarity, purple_rarity, blue_rarity = \
        getting_chances(gold_items, red_items, pink_items, purple_items, blue_items)

    gold_item_rarity = gold_rarity / len(gold_items) if len(gold_items) > 0 else 0
    red_item_rarity = red_rarity / len(red_items) if len(red_items) > 0 else 0
    pink_item_rarity = pink_rarity / len(pink_items) if len(pink_items) > 0 else 0
    purple_item_rarity = purple_rarity / len(purple_items) if len(
        purple_items) > 0 else 0
    blue_item_rarity = blue_rarity / len(blue_items) if len(blue_items) > 0 else 0

    page.cell(column=cell_x, row=cell_y,
              value=f'{case_num}. {case_name} ({case_price} руб.)')
    cell_name_x, cell_name_y = cell_x, cell_y
    cell_x += 1
    case_info = list()
    dispersion_list = list()

    if key_flag:
        key_price = CONST_KEY_PRICE
    else:
        key_price = 0

    if blue_items:
        excel_decoration_for_rarity(page, cell_x, cell_y, 17)
        page.cell(column=cell_x, row=cell_y,
                  value=f'blue ({round(blue_rarity * 100, 2)}%)')
        cell_y += 1

        blue_items_info = list()

        for item in all_items:
            if len(item.xpath('.//a/@title')) > 0:
                if item.xpath('.//a/@title')[0] == 'All Mil-Spec Skins':
                    item_load(item, cell_x, cell_y, blue_item_rarity, case_price,
                              blue_items_info, page, '0070c0',
                              dispersion_list, key_price)
                    cell_y += 2

        if len(blue_items_info) > 0:
            if len(blue_items_info[0]) == 8:
                load_info_for_case(blue_items_info, page, cell_x, cell_y, case_info,
                                   '0070c0', blue_item_rarity)
                cell_y += 2
                cell_x = 3

    if purple_items:
        excel_decoration_for_rarity(page, cell_x, cell_y, 17)
        page.cell(column=cell_x, row=cell_y,
                  value=f'purple ({round(purple_rarity * 100, 2)}%)')
        cell_y += 1

        purple_items_info = list()

        for item in all_items:
            if len(item.xpath('.//a/@title')) > 0:
                if item.xpath('.//a/@title')[0] == 'All Restricted Skins':
                    item_load(item, cell_x, cell_y, purple_item_rarity, case_price,
                              purple_items_info, page,
                              '7030a0', dispersion_list, key_price)
                    cell_y += 2

        if len(purple_items_info) > 0:
            if len(purple_items_info[0]) == 8:
                load_info_for_case(purple_items_info, page, cell_x, cell_y, case_info,
                                   '7030a0', purple_item_rarity)
                cell_y += 2
                cell_x = 3

    if pink_items:
        excel_decoration_for_rarity(page, cell_x, cell_y, 17)
        page.cell(column=cell_x, row=cell_y,
                  value=f'pink ({round(pink_rarity * 100, 2)}%)')
        cell_y += 1

        pink_items_info = list()

        for item in all_items:
            if len(item.xpath('.//a/@title')) > 0:
                if item.xpath('.//a/@title')[0] == 'All Classified Skins':
                    item_load(item, cell_x, cell_y, pink_item_rarity, case_price,
                              pink_items_info, page, 'cc00ff',
                              dispersion_list, key_price)
                    cell_y += 2

        if len(pink_items_info) > 0:
            if len(pink_items_info[0]) == 8:
                load_info_for_case(pink_items_info, page, cell_x, cell_y, case_info,
                                   'cc00ff', pink_item_rarity)
                cell_y += 2
                cell_x = 3

    if red_items:
        excel_decoration_for_rarity(page, cell_x, cell_y, 17)
        page.cell(column=cell_x, row=cell_y,
                  value=f'red ({round(red_rarity * 100, 2)}%)')
        cell_y += 1

        red_items_info = list()

        for item in all_items:
            if len(item.xpath('.//a/@title')) > 0:
                if item.xpath('.//a/@title')[0] == 'All Covert Skins':
                    item_load(item, cell_x, cell_y, red_item_rarity, case_price,
                              red_items_info, page, 'cc0000',
                              dispersion_list, key_price)
                    cell_y += 2

        if len(red_items_info) > 0:
            if len(red_items_info[0]) == 8:
                load_info_for_case(red_items_info, page, cell_x, cell_y, case_info,
                                   'cc0000', red_item_rarity)
                cell_y += 2
                cell_x = 3

    s = 0
    for i in dispersion_list:
        s += i[1]

    print('Сумма вероятностей:', s, round(s, 6))
    print()

    if round(s, 6) != 0.997437 and round(s, 6) != 1:
        print(
            'ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ERROR ')
        a = input()

    final_case_score = 0
    average_dispersion_cost_for_case = 0

    for item_info in case_info:
        final_case_score += item_info[3]
        average_dispersion_cost_for_case += item_info[4]

    disp = dispersion_score(dispersion_list, average_dispersion_cost_for_case,
                            case_price)

    case_info.append(final_case_score)
    case_info.append(case_name)
    case_info.append(disp)
    case_info.append(case_price)

    page.cell(column=cell_name_x, row=cell_name_y + 1,
              value=f'Итоговый счет кейса: {final_case_score}')
    page.cell(column=cell_name_x, row=cell_name_y + 2, value=f'Дисперсия кейса: {disp}')

    print(case_info)

    return case_info, cell_y, case_num


def cases_load(page_name, key_flag, url):
    cases_url = url

    r = requests.get(cases_url)

    with open('cases_page.html', 'w', encoding="utf-8") as output_file:
        output_file.write(r.text)

    cases_page_html = read_file('cases_page.html')

    tree = html.fromstring(cases_page_html)
    cases_links_list = tree.xpath('//div[@class = "well result-box nomargin"]/a/@href')
    cases_names_list = tree.xpath(
        '//div[@class = "well result-box nomargin"]/a/h4/text()')
    cases_prices_list = tree.xpath('//div[@class = "well result-box nomargin"]/a'
                                   '/div[@class = "price margin-top-sm"]/p/text()')

    cases_info = dict()
    for i in range(len(cases_names_list)):
        cases_info[cases_names_list[i]] = [cases_links_list[i],
                                           float(cases_prices_list[i].replace('pуб.',
                                                                              '').replace(
                                               ' ', '')
                                                 .replace(',', '.'))]

    wb = load_workbook('cases.xlsx')

    try:
        ws = wb[page_name]
    except KeyError:
        ws = wb.create_sheet(page_name)

    try:
        rating_page = wb[f'Рейтинг ({page_name})']
    except KeyError:
        rating_page = wb.create_sheet(f'Рейтинг ({page_name})')

    cell_x = 2
    cell_y = 2
    ss = 0

    cases_full_info = list()

    case_num = 1
    for case_name, case_info in cases_info.items():
        if case_name == 'X-Ray P250 Package':
            continue
            # case_full_info, cell_y, case_num = case_load(case_name, case_info[0], case_info[1], ws, cell_x, cell_y,
            #                                              case_num, False)
        # if ss in range(0, 34):
        #     ss += 1
        #     continue

        # if ss == 5:
        #     break

        if case_name == 'Anubis Collection Package':
            continue

        for cells_tuple in ws['B2':'S2']:
            thins = Side(border_style="thin", color="000000")
            for cell in cells_tuple:
                ws.cell(column=cell_x, row=cell_y, value=cell.value).border = Border(
                    left=thins, right=thins,
                    bottom=thins, top=thins)
                ws.cell(column=cell_x, row=cell_y,
                        value=cell.value).alignment = Alignment(horizontal='center')
                cell_x += 1

        cell_x = 2
        cell_y += 1

        case_full_info, cell_y, case_num = case_load(case_name, case_info[0],
                                                     case_info[1], ws, cell_x, cell_y,
                                                     case_num, key_flag)

        cases_full_info.append(case_full_info)

        print(
            '---------------------------------------------------------------------------')
        print('Информация о кейсе: ', case_full_info)
        print(
            '---------------------------------------------------------------------------')
        print()

        ss += 1
        case_num += 1

    cases_rating(cases_full_info, rating_page)

    print('Окупаемость:')
    for case_info in sorted(cases_full_info, key=lambda x: x[4], reverse=True):
        print(case_info[5])
    print()

    print('Дисперсия:')
    for case_info in sorted(cases_full_info, key=lambda x: x[6], reverse=True):
        print(case_info[5])
    print()

    print('Синее:')
    for case_info in sorted(cases_full_info, key=lambda x: x[0][2], reverse=True):
        print(case_info[5])
    print()

    print('Фиолетовое:')
    for case_info in sorted(cases_full_info, key=lambda x: x[1][2], reverse=True):
        print(case_info[5])
    print()

    print('Розовое:')
    for case_info in sorted(cases_full_info, key=lambda x: x[2][2], reverse=True):
        print(case_info[5])
    print()

    print('Красное:')
    for case_info in sorted(cases_full_info, key=lambda x: x[3][2], reverse=True):
        print(case_info[5])
    print()

    print('Информация о кейсах: ',
          sorted(cases_full_info, key=lambda x: x[4], reverse=True))
    print()

    wb.save('cases.xlsx')


def cases_rating(cases_full_info, rating_page):
    final_rating = dict()

    cell_x = 2
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[7]):

        value = case[7]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({value} руб.)')
        cases_full_info[cases_full_info.index(case)].append(case_num)
        cell_y += 1

    print('!!!!', cases_full_info)

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[4], reverse=True):

        value = case[4]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({round(value, 4)})')
        final_rating[case[5]] = case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[6], reverse=True):

        value = case[6]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({round(value, 2)}), ('
                               f'Рейтинг по цене: {case[8]})')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[0][2], reverse=True):

        value = case[0][2]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[1][2], reverse=True):

        value = case[1][2]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[2][2], reverse=True):

        value = case[2][2]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: x[3][2], reverse=True):

        value = case[3][2]

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y,
                         value=f'{case_num}. {case[5]} ({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: (x[0][0] + x[0][1]) / 2,
                       reverse=True):

        value = (case[0][0] + case[0][1]) / 2

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y, value=f'{case_num}. {case[5]} '
                                                          f'({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: (x[1][0] + x[1][1]) / 2,
                       reverse=True):

        value = (case[1][0] + case[1][1]) / 2

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y, value=f'{case_num}. {case[5]} '
                                                          f'({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: (x[2][0] + x[2][1]) / 2,
                       reverse=True):

        value = (case[2][0] + case[2][1]) / 2

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y, value=f'{case_num}. {case[5]} '
                                                          f'({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: (x[3][0] + x[3][1]) / 2,
                       reverse=True):

        value = (case[3][0] + case[3][1]) / 2

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y, value=f'{case_num}. {case[5]} '
                                                          f'({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for case in sorted(cases_full_info, key=lambda x: (x[3][0] + x[3][1] + x[2][0] +
                                                       x[2][1] + x[1][0] + x[1][1] +
                                                       x[0][0] + x[0][1]) / 8,
                       reverse=True):
        value = (case[3][0] + case[3][1] + case[2][0] + case[2][1] + case[1][0] +
                 case[1][1] + case[0][0] +
                 case[0][1]) / 8

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y, value=f'{case_num}. {case[5]} '
                                                          f'({round(value * 100, 2)}%)')
        final_rating[case[5]] += case_num
        cell_y += 1

    cell_x += 1
    cell_y = 3
    case_num = 0
    last_value = 0

    for key, value in sorted(final_rating.items(), key=lambda x: x[1]):

        if value != last_value and case_num != 0:
            case_num += 1

        elif case_num == 0:
            case_num += 1

        last_value = value

        rating_page.cell(column=cell_x, row=cell_y, value=f'{case_num}. {key} ({value})')

        cell_y += 1


if __name__ == "__main__":
    # cases_load('Капсулы', False, 'https://csgostash.com/containers/sticker-capsules')
    cases_load('Кейсы', True, 'https://csgostash.com/containers/skin-cases')
